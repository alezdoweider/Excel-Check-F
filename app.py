import streamlit as st
import pandas as pd
import openpyxl
import io

# Configurar la página en modo ancho
st.set_page_config(page_title="Gestor de Casos (BlueStars)", layout="wide")

st.markdown("""
<style>
body { background-color: black; color: white; }
.stApp { background-color: black; color: white; }
[data-testid="stSidebar"] { background-color: #222222; }
h1, h2, h3, h4, h5, h6, label, p, div, span {
  color: white !important;
}
.table-cell {
    border: 1px solid white;
    padding: 5px;
    text-align: center;
    word-wrap: break-word;
    white-space: normal;
}
.table-header {
    border: 2px solid white;
    padding: 8px;
    font-weight: bold;
    background-color: #444;
    text-align: center;
    word-wrap: break-word;
    white-space: normal;
}
</style>
""", unsafe_allow_html=True)

def main():
    st.title("Gestor de Casos (BlueStars)")

    uploaded_file = st.file_uploader("Sube el archivo Excel (.xlsm)", type=["xlsm"])
    if uploaded_file:
        try:
            # Cargar el archivo en memoria para modificarlo
            file_stream = io.BytesIO(uploaded_file.getvalue())
            wb = openpyxl.load_workbook(file_stream, keep_vba=True)
            
            if 'ARMADRE' not in wb.sheetnames:
                st.error("La hoja 'ARMADRE' no se encuentra en el archivo.")
                return
            
            ws = wb['ARMADRE']
            df = pd.DataFrame(ws.values)
            st.success("Archivo cargado correctamente")

            df['CASO'] = df.iloc[:, 16].astype(str).apply(lambda x: x.split('-')[0])
            df['NUNC'] = df.iloc[:, 16].apply(lambda x: str(x).split('-')[1] if '-' in str(x) else '')
            df['NOMBRE'] = df.iloc[:, 10].astype(str)
            df['ID'] = df.iloc[:, 4].astype(str)
            df['NRO ID'] = pd.to_numeric(df.iloc[:, 5], errors='coerce')
            df['TIPO DE EMP'] = df.iloc[:, 7].astype(str)

            lista_casos = df['CASO'].dropna().unique().tolist()
            if not lista_casos:
                st.warning("No se encontraron valores en la columna Q para generar la lista de casos.")
                return

            caso_seleccionado = st.selectbox("Selecciona un CASO:", lista_casos)
            if caso_seleccionado:
                st.subheader(f"Información del CASO: {caso_seleccionado}")
                df_filtrado = df[df['CASO'] == caso_seleccionado].copy()
                df_filtrado.reset_index(drop=True, inplace=True)
                
                envase_options = ["TTG", "TTR", "TTL", "TTV", "FP", "BP"]
                
                # Asegurar que la columna 'TIPO DE ENVASE' exista antes de usarla
                if 'TIPO DE ENVASE' not in df_filtrado.columns:
                    df_filtrado['TIPO DE ENVASE'] = ""
                
                # Convertir la columna 'TIPO DE ENVASE' en un dropdown dentro de la tabla
                edited_df = st.data_editor(
                    df_filtrado,
                    column_config={
                        "TIPO DE ENVASE": st.column_config.SelectboxColumn(
                            "TIPO DE ENVASE",
                            options=envase_options
                        )
                    },
                    use_container_width=True
                )
                
                columnas_finales = ['CASO', 'NUNC', 'NOMBRE', 'ID', 'NRO ID', 'TIPO DE EMP', 'TIPO DE ENVASE']
                st.write("### Resultado final con Tipo de Envase seleccionado:")
                st.dataframe(edited_df[columnas_finales], use_container_width=True)

                # Ajustar ancho de la columna 'NRO ID'
                if 'ARMADRE' in wb.sheetnames:
                    ws_armadre = wb['ARMADRE']
                    ws_armadre.column_dimensions['F'].width = 30

                # Entradas de usuario adicionales
                anio = st.text_input("Ingrese el AÑO:")
                mes = st.text_input("Ingrese el MES:")
                dia = st.text_input("Ingrese el DÍA:")
                custodio = st.text_input("Ingrese el CUSTODIO:")

                if st.button("Guardar Información"):
                    if 'HT' in wb.sheetnames:
                        ws_ht = wb['HT']
                        ws_ht['AA7'] = anio
                        ws_ht['AE7'] = mes
                        ws_ht['AG7'] = dia
                        ws_ht['AE11'] = custodio
                    if 'LCH' in wb.sheetnames:
                        ws_lch = wb['LCH']
                        ws_lch['H9'] = custodio
                    
                    # Guardar los cambios en el archivo original
                    wb.save(file_stream)
                    st.success("Información guardada correctamente en el archivo.")
        except Exception as e:
            st.error(f"Error al leer el archivo: {e}")

if __name__ == "__main__":
    main()
